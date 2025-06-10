"""Export functionality for search results."""

import csv
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger = logging.getLogger(__name__)
    logger.warning("openpyxl not available, Excel export disabled")

from chemscreen.models import SearchResult, QualityMetrics, BatchSearchSession
from chemscreen.config import get_config, Config

logger = logging.getLogger(__name__)


class ExportManager:
    """Manages export of search results to various formats."""

    def __init__(
        self, export_dir: Optional[Path] = None, config: Optional[Config] = None
    ):
        """
        Initialize export manager.

        Args:
            export_dir: Directory for export files (uses config if None)
            config: Configuration instance (uses global if None)
        """
        self.config = config or get_config()
        self.export_dir = export_dir or self.config.exports_dir
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def export_to_csv(
        self,
        results: list[tuple[SearchResult, QualityMetrics]],
        session: BatchSearchSession,
        filename: Optional[str] = None,
        include_abstracts: bool = False,
    ) -> Path:
        """
        Export results to CSV format.

        Args:
            results: List of (SearchResult, QualityMetrics) tuples
            session: Batch search session
            filename: Output filename (generated if not provided)
            include_abstracts: Include publication abstracts

        Returns:
            Path to exported file
        """
        if not filename:
            filename = f"chemscreen_export_{session.batch_id}.csv"

        filepath = self.export_dir / filename

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            # Define fields
            fields = [
                "Chemical Name",
                "CAS Number",
                "Total Publications",
                "Review Articles",
                "Recent Publications (3yr)",
                "Quality Score",
                "Publication Trend",
                "Has Recent Review",
                "Search Status",
                "Error Message",
            ]

            if include_abstracts:
                fields.extend(
                    ["PMID", "Title", "Authors", "Journal", "Year", "Abstract"]
                )

            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()

            # Write data
            for result, metrics in results:
                base_row = {
                    "Chemical Name": result.chemical.name,
                    "CAS Number": result.chemical.cas_number or "",
                    "Total Publications": metrics.total_publications,
                    "Review Articles": metrics.review_count,
                    "Recent Publications (3yr)": metrics.recent_publications,
                    "Quality Score": metrics.quality_score,
                    "Publication Trend": metrics.publication_trend,
                    "Has Recent Review": "Yes" if metrics.has_recent_review else "No",
                    "Search Status": "Failed" if result.error else "Success",
                    "Error Message": result.error or "",
                }

                if include_abstracts and result.publications:
                    # Write one row per publication
                    for pub in result.publications:
                        row = base_row.copy()
                        row.update(
                            {
                                "PMID": pub.pmid,
                                "Title": pub.title,
                                "Authors": "; ".join(pub.authors),
                                "Journal": pub.journal or "",
                                "Year": pub.year or "",
                                "Abstract": pub.abstract or "",
                            }
                        )
                        writer.writerow(row)
                else:
                    # Write summary row only
                    writer.writerow(base_row)

        logger.info(f"Exported CSV to {filepath}")
        return filepath

    def export_to_excel(
        self,
        results: list[tuple[SearchResult, QualityMetrics]],
        session: BatchSearchSession,
        filename: Optional[str] = None,
        include_abstracts: bool = False,
    ) -> Optional[Path]:
        """
        Export results to Excel format with multiple sheets.

        Args:
            results: List of (SearchResult, QualityMetrics) tuples
            session: Batch search session
            filename: Output filename (generated if not provided)
            include_abstracts: Include publication abstracts

        Returns:
            Path to exported file or None if Excel not available
        """
        if not EXCEL_AVAILABLE:
            logger.error("Excel export not available, openpyxl not installed")
            return None

        if not filename:
            filename = f"chemscreen_export_{session.batch_id}.xlsx"

        filepath = self.export_dir / filename

        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        if wb.active is not None:
            wb.remove(wb.active)

        # Create sheets
        self._create_summary_sheet(wb, results, session)
        self._create_detailed_sheet(wb, results, include_abstracts)
        self._create_metadata_sheet(wb, session)

        # Save workbook
        wb.save(filepath)
        logger.info(f"Exported Excel to {filepath}")

        return filepath

    def _create_summary_sheet(
        self,
        wb: Any,
        results: list[tuple[SearchResult, QualityMetrics]],
        session: BatchSearchSession,
    ) -> None:
        """Create summary sheet in Excel workbook."""
        ws = wb.create_sheet("Summary")

        # Headers
        headers = [
            "Chemical Name",
            "CAS Number",
            "Total Publications",
            "Review Articles",
            "Recent Publications",
            "Quality Score",
            "Trend",
            "Priority",
        ]

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="0066CC", end_color="0066CC", fill_type="solid"
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row, (result, metrics) in enumerate(results, 2):
            ws.cell(row=row, column=1, value=result.chemical.name)
            ws.cell(row=row, column=2, value=result.chemical.cas_number or "")
            ws.cell(row=row, column=3, value=metrics.total_publications)
            ws.cell(row=row, column=4, value=metrics.review_count)
            ws.cell(row=row, column=5, value=metrics.recent_publications)
            ws.cell(row=row, column=6, value=metrics.quality_score)
            ws.cell(row=row, column=7, value=metrics.publication_trend)

            # Priority based on quality score
            if metrics.quality_score >= 70:
                priority = "High"
                color = "00AA00"  # Green
            elif metrics.quality_score >= 40:
                priority = "Medium"
                color = "FFA500"  # Orange
            else:
                priority = "Low"
                color = "FF0000"  # Red

            priority_cell = ws.cell(row=row, column=8, value=priority)
            priority_cell.font = Font(color=color, bold=True)

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_detailed_sheet(
        self,
        wb: Any,
        results: list[tuple[SearchResult, QualityMetrics]],
        include_abstracts: bool,
    ) -> None:
        """Create detailed results sheet in Excel workbook."""
        ws = wb.create_sheet("Detailed Results")

        # Headers
        headers = [
            "Chemical Name",
            "CAS Number",
            "PMID",
            "Title",
            "Authors",
            "Journal",
            "Year",
            "Type",
        ]

        if include_abstracts:
            headers.append("Abstract")

        # Style headers
        header_font = Font(bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font

        # Write data
        row_num = 2
        for result, metrics in results:
            if result.publications:
                for pub in result.publications:
                    ws.cell(row=row_num, column=1, value=result.chemical.name)
                    ws.cell(
                        row=row_num, column=2, value=result.chemical.cas_number or ""
                    )
                    ws.cell(row=row_num, column=3, value=pub.pmid)
                    ws.cell(row=row_num, column=4, value=pub.title)
                    ws.cell(
                        row=row_num,
                        column=5,
                        value="; ".join(pub.authors[:3])
                        + ("..." if len(pub.authors) > 3 else ""),
                    )
                    ws.cell(row=row_num, column=6, value=pub.journal or "")
                    ws.cell(row=row_num, column=7, value=pub.year or "")
                    ws.cell(
                        row=row_num,
                        column=8,
                        value="Review" if pub.is_review else "Research",
                    )

                    if include_abstracts:
                        ws.cell(row=row_num, column=9, value=pub.abstract or "")

                    row_num += 1
            else:
                # Chemical with no results
                ws.cell(row=row_num, column=1, value=result.chemical.name)
                ws.cell(row=row_num, column=2, value=result.chemical.cas_number or "")
                ws.cell(
                    row=row_num,
                    column=3,
                    value="No results"
                    if not result.error
                    else f"Error: {result.error}",
                )
                row_num += 1

    def _create_metadata_sheet(self, wb: Any, session: BatchSearchSession) -> None:
        """Create metadata sheet in Excel workbook."""
        ws = wb.create_sheet("Search Metadata")

        # Write metadata
        metadata = [
            ("Batch ID", session.batch_id),
            ("Search Date", session.created_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Chemicals", len(session.chemicals)),
            ("Date Range (years)", session.parameters.date_range_years),
            ("Max Results per Chemical", session.parameters.max_results),
            ("Include Reviews", "Yes" if session.parameters.include_reviews else "No"),
            ("Cache Used", "Yes" if session.parameters.use_cache else "No"),
            ("Export Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        for row, (label, value) in enumerate(metadata, 1):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value))

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

    def export_to_json(
        self,
        results: list[tuple[SearchResult, QualityMetrics]],
        session: BatchSearchSession,
        filename: Optional[str] = None,
    ) -> Path:
        """
        Export results to JSON format.

        Args:
            results: List of (SearchResult, QualityMetrics) tuples
            session: Batch search session
            filename: Output filename (generated if not provided)

        Returns:
            Path to exported file
        """
        if not filename:
            filename = f"chemscreen_export_{session.batch_id}.json"

        filepath = self.export_dir / filename

        # Build export data structure
        export_data = {
            "metadata": {
                "batch_id": session.batch_id,
                "search_date": session.created_at.isoformat(),
                "total_chemicals": len(session.chemicals),
                "parameters": {
                    "date_range_years": session.parameters.date_range_years,
                    "max_results": session.parameters.max_results,
                    "include_reviews": session.parameters.include_reviews,
                    "use_cache": session.parameters.use_cache,
                },
            },
            "results": [],
        }

        # Add results
        for result, metrics in results:
            chemical_data = {
                "chemical": {
                    "name": result.chemical.name,
                    "cas_number": result.chemical.cas_number,
                    "validated": result.chemical.validated,
                },
                "metrics": {
                    "total_publications": metrics.total_publications,
                    "review_count": metrics.review_count,
                    "recent_publications": metrics.recent_publications,
                    "quality_score": metrics.quality_score,
                    "publication_trend": metrics.publication_trend,
                    "has_recent_review": metrics.has_recent_review,
                },
                "search_info": {
                    "status": "success" if not result.error else "failed",
                    "error": result.error,
                    "search_time_seconds": result.search_time_seconds,
                    "from_cache": result.from_cache,
                },
                "publications": [
                    {
                        "pmid": pub.pmid,
                        "title": pub.title,
                        "authors": pub.authors,
                        "journal": pub.journal,
                        "year": pub.year,
                        "is_review": pub.is_review,
                    }
                    for pub in result.publications
                ],
            }
            export_data["results"].append(chemical_data)  # type: ignore[attr-defined]

        # Write JSON
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)

        logger.info(f"Exported JSON to {filepath}")
        return filepath
